import streamlit as st
import plotly.express as px
import pandas as pd
import logging
import shutil
import yaml
import os
import json
import io
import zipfile

from autogluon.tabular import TabularPredictor
from openpyxl.styles import PatternFill

# Импорт функций обработки данных, feature engineering и prediction
from src.data.data_processing import (
    load_data,
    show_dataset_stats
)
from src.features.feature_engineering import (
    fill_missing_values
)
from src.models.prediction import (
    predict_tabular
)
from src.utils.utils import (
    setup_logger,
    read_logs,
    LOG_FILE
)
from src.help_page import show_help_page

CONFIG_PATH = "config/config.yaml"
MODEL_DIR = "AutogluonModels/TabularModel"
MODEL_INFO_FILE = "model_info.json"

def load_config(path: str):
    """Загружает YAML конфигурацию (METRICS_DICT, AG_MODELS, PRESETS_LIST)."""
    if not os.path.exists(path):
        raise FileNotFoundError(f"Конфигурационный файл {path} не найден.")
    with open(path, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f)
    metrics_dict = data.get("metrics_dict", {})
    ag_models = data.get("ag_models", {})
    presets_list = data.get("presets_list", [])
    logging.info(f"Конфигурация загружена из: {path}")
    return metrics_dict, ag_models, presets_list

METRICS_DICT, AG_MODELS, PRESETS_LIST = load_config(CONFIG_PATH)

def save_model_metadata(
    целевая_колонка,
    тип_задачи,
    метрика_оценки,
    метод_заполнения_пропусков,
    группировочные_колонки_для_заполнения,
    пресеты,
    выбранные_модели
):
    """Сохраняет все настройки (колонки, тип задачи, метрика и т.д.) в JSON (для автозагрузки)."""
    os.makedirs(MODEL_DIR, exist_ok=True)
    info_dict = {
        "целевая_колонка": целевая_колонка,
        "тип_задачи": тип_задачи,
        "метрика_оценки": метрика_оценки,
        "метод_заполнения_пропусков": метод_заполнения_пропусков,
        "группировочные_колонки_для_заполнения": группировочные_колонки_для_заполнения,
        "пресеты": пресеты,
        "выбранные_модели": выбранные_модели,
    }
    path_json = os.path.join(MODEL_DIR, MODEL_INFO_FILE)
    with open(path_json, "w", encoding="utf-8") as f:
        json.dump(info_dict, f, ensure_ascii=False, indent=2)
    logging.info(f"Метаданные модели сохранены: {path_json}")

def load_model_metadata():
    """Загружает сохраненные настройки из model_info.json, если есть."""
    path_json = os.path.join(MODEL_DIR, MODEL_INFO_FILE)
    if not os.path.exists(path_json):
        logging.info(f"Файл метаданных не найден: {path_json}")
        return None
    try:
        with open(path_json, "r", encoding="utf-8") as f:
            info = json.load(f)
        return info
    except Exception as e:
        logging.error(f"Ошибка при загрузке метаданных: {e}", exc_info=True)
        return None

def try_load_existing_model():
    """Пытается загрузить TabularPredictor и метаданные, если есть в MODEL_DIR."""
    if not os.path.exists(MODEL_DIR):
        logging.info(f"Папка с моделью {MODEL_DIR} не найдена, пропускаем автозагрузку.")
        return
    try:
        loaded_predictor = TabularPredictor.load(MODEL_DIR)
        st.session_state["predictor"] = loaded_predictor
        st.info(f"Загружена ранее обученная модель из {MODEL_DIR}")

        meta = load_model_metadata()
        if meta:
            st.session_state["tgt_col_key"] = meta.get("целевая_колонка", "<нет>")
            st.session_state["problem_type_key"] = meta.get("тип_задачи", "auto")
            st.session_state["eval_metric_key"] = meta.get("метрика_оценки", "auto")
            st.session_state["fill_method_key"] = meta.get("метод_заполнения_пропусков", "None")
            st.session_state["group_cols_for_fill_key"] = meta.get("группировочные_колонки_для_заполнения", [])
            st.session_state["presets_key"] = meta.get("пресеты", "medium_quality")
            st.session_state["models_key"] = meta.get("выбранные_модели", ["* (all)"])

            st.info("Настройки восстановлены из model_info.json")
    except Exception as e:
        st.warning(f"Не удалось загрузить модель из {MODEL_DIR}: {e}")

def display_fit_summary(fit_summary):
    """Отображает резюме обучения в удобном виде."""
    if fit_summary is None:
        st.warning("Резюме обучения отсутствует.")
        return

    st.subheader("Общая информация")
    st.write(f"- Тип проблемы: {fit_summary.get('problem_type', 'Н/Д')}")
    st.write(f"- Метрика оценки: {fit_summary.get('eval_metric', 'Н/Д')}")
    st.write(f"- Лучшая модель: {fit_summary.get('model_best', 'Н/Д')}")
    st.write(f"- Количество классов: {fit_summary.get('num_classes', 'Н/Д')}")
    st.write(f"- Количество фолдов: {fit_summary.get('num_bag_folds', 'Н/Д')}")
    st.write(f"- Максимальный уровень стекинга: {fit_summary.get('max_stack_level', 'Н/Д')}")

    st.subheader("Производительность моделей")
    perf = fit_summary.get('model_performance', {})
    if perf:
        df_perf = pd.DataFrame.from_dict(perf, orient='index', columns=['Score'])
        st.dataframe(df_perf.sort_values(by='Score', ascending=False))
    else:
        st.write("Нет информации о производительности.")

    st.subheader("Время обучения")
    fit_times = fit_summary.get('model_fit_times', {})
    if fit_times:
        df_fit = pd.DataFrame.from_dict(fit_times, orient='index', columns=['FitTime(sec)'])
        st.dataframe(df_fit.sort_values(by='FitTime(sec)', ascending=True))
    else:
        st.write("Нет информации о времени обучения.")

    st.subheader("Время прогнозирования")
    pred_times = fit_summary.get('model_pred_times', {})
    if pred_times:
        df_pred = pd.DataFrame.from_dict(pred_times, orient='index', columns=['PredTime(sec)'])
        st.dataframe(df_pred.sort_values(by='PredTime(sec)', ascending=True))
    else:
        st.write("Нет информации о времени прогноза.")

    st.subheader("Гиперпараметры моделей")
    model_hps = fit_summary.get('model_hyperparams', {})
    if model_hps:
        st.json(model_hps)
    else:
        st.write("Нет гиперпараметров.")

def extract_ensemble_info(predictor, best_model_name: str):
    """Если лучшая модель — WeightedEnsemble, получаем состав (модель/вес)."""
    if not best_model_name.startswith("WeightedEnsemble"):
        return None
    info_dict = predictor.info()
    model_info = info_dict.get("model_info", {})
    best_info = model_info.get(best_model_name, {})
    child_info = best_info.get("children_info", {})
    weights = child_info.get("child_weights", None)
    child_models = child_info.get("child_model_names", None)
    if not weights or not child_models:
        return None  # Возможно другая версия AutoGluon
    data_rows = []
    for m, w in zip(child_models, weights):
        data_rows.append({"Model": m, "Weight": w})
    return pd.DataFrame(data_rows)

def main():
    setup_logger()
    logging.info("=== Запуск приложения (Tabular) ===")

    # Пытаемся автозагрузить модель (если есть)
    if "predictor" not in st.session_state or st.session_state["predictor"] is None:
        try_load_existing_model()

    pages = ["Главная", "Help"]
    choice = st.sidebar.selectbox("Навигация", pages, key="page_choice")

    if choice == "Help":
        show_help_page()
        return

    st.title("Приложение AutoGluon (Tabular)")

    # Инициализация session_state
    for key_ in ["df","df_predict","predictor","leaderboard","predictions","fit_summary","feature_importance","ensemble_info"]:
        if key_ not in st.session_state:
            st.session_state[key_] = None

    # ---------- Очистка логов ----------
    st.sidebar.header("Очистка логов")
    clear_logs_input = st.sidebar.text_input("Введите 'delete' чтобы очистить логи:")
    if st.sidebar.button("Очистить логи"):
        if clear_logs_input.strip().lower() == "delete":
            logger = logging.getLogger()
            for handler in logger.handlers[:]:
                if hasattr(handler, 'baseFilename') and os.path.abspath(handler.baseFilename) == os.path.abspath(LOG_FILE):
                    handler.close()
                    logger.removeHandler(handler)

            try:
                if os.path.exists(LOG_FILE):
                    os.remove(LOG_FILE)
                    st.warning("Логи успешно удалены!")
                else:
                    st.info("Файл логов не найден, нечего удалять.")
            except Exception as e:
                st.error(f"Ошибка при удалении лог-файла: {e}")

            # Пересоздаём пустой лог-файл
            with open(LOG_FILE, 'w', encoding='utf-8') as f:
                f.write("")
            # Добавим новый handler
            file_handler = logging.FileHandler(LOG_FILE, mode='a', encoding='utf-8')
            formatter = logging.Formatter(
                "%(asctime)s [%(levelname)s] %(module)s.%(funcName)s - %(message)s",
                datefmt="%Y-%m-%d %H:%M:%S"
            )
            file_handler.setFormatter(formatter)
            logger.addHandler(file_handler)
            logger.info("Создан новый log-файл после очистки.")
        else:
            st.warning("Вы ввели неверное слово — логи не очищены.")

    # ========== (1) Загрузка данных ==========
    st.sidebar.header("1. Загрузка данных")
    train_file = st.sidebar.file_uploader("Тренировочные данные (обязательно)", type=["csv", "xls", "xlsx"], key="train_file_uploader")
    predict_file = st.sidebar.file_uploader("Данные для прогнозирования (обязательно)", type=["csv", "xls", "xlsx"], key="predict_file_uploader")

    if st.sidebar.button("Загрузить данные", key="load_data_btn"):
        if not train_file:
            st.error("Файл Train обязателен!")
        elif not predict_file:
            st.error("Файл для прогнозирования обязателен!")
        else:
            try:
                df_train = load_data(train_file)
                st.session_state["df"] = df_train
                st.success("Train-файл загружен!")
                st.dataframe(df_train.head())

                st.subheader("Статистика Train")
                show_dataset_stats(df_train)

                df_predict = load_data(predict_file)
                st.session_state["df_predict"] = df_predict
                st.success("Файл для прогнозирования загружен!")
                st.dataframe(df_predict.head())

                st.subheader("Статистика Data для прогнозов")
                show_dataset_stats(df_predict)
            except Exception as e:
                st.error(f"Ошибка загрузки: {e}")

    # ========== (2) Настройка колонок ==========
    st.sidebar.header("2. Настройка колонок")
    df_current = st.session_state["df"]
    all_cols = list(df_current.columns) if df_current is not None else []

    # Защита от «ValueError: ... not in iterable»
    if "tgt_col_key" not in st.session_state:
        st.session_state["tgt_col_key"] = "<нет>"
    else:
        if st.session_state["tgt_col_key"] not in ["<нет>"] + all_cols:
            st.session_state["tgt_col_key"] = "<нет>"

    tgt_col = st.sidebar.selectbox("Целевая колонка", ["<нет>"] + all_cols, key="tgt_col_key")

    problem_type_options = ["auto", "binary", "multiclass", "regression"]
    if "problem_type_key" not in st.session_state:
        st.session_state["problem_type_key"] = "auto"
    problem_type = st.sidebar.selectbox("Тип задачи", problem_type_options, index=0, key="problem_type_key")

    eval_metric_options = ["auto"] + list(METRICS_DICT.keys())
    if "eval_metric_key" not in st.session_state:
        st.session_state["eval_metric_key"] = "auto"
    eval_metric = st.sidebar.selectbox("Метрика оценки", eval_metric_options, index=0, key="eval_metric_key")

    # ========== (3) Обработка пропусков ==========
    st.sidebar.header("3. Пропущенные значения")
    fill_options = ["None", "Constant=0", "Mean", "Median", "Mode"]
    if "fill_method_key" not in st.session_state:
        st.session_state["fill_method_key"] = "None"
    fill_method = st.sidebar.selectbox("Метод заполнения пропусков", fill_options, key="fill_method_key")

    group_cols_for_fill = []

    # ========== (4) Настройки модели ==========
    st.sidebar.header("4. Настройки модели и обучения")
    model_keys = list(AG_MODELS.keys())
    model_choices = ["* (all)"] + model_keys
    if "models_key" not in st.session_state:
        st.session_state["models_key"] = ["* (all)"]
    chosen_models = st.sidebar.multiselect(
        "Модели AutoGluon",
        model_choices,
        default=st.session_state["models_key"],
        key="models_key"
    )

    if "presets_key" not in st.session_state:
        st.session_state["presets_key"] = "medium_quality"
    presets = st.sidebar.selectbox(
        "Пресеты",
        PRESETS_LIST,
        index=PRESETS_LIST.index(st.session_state["presets_key"]) if st.session_state["presets_key"] in PRESETS_LIST else 0,
        key="presets_key"
    )

    if "time_limit_key" not in st.session_state:
        st.session_state["time_limit_key"] = 60
    time_limit = st.sidebar.number_input("Лимит времени обучения (сек)", 10, 36000, st.session_state["time_limit_key"], key="time_limit_key")

    # ========== (5) Обучение модели ==========
    st.sidebar.header("5. Обучение модели")
    if "auto_predict_save_checkbox" not in st.session_state:
        st.session_state["auto_predict_save_checkbox"] = False
    auto_predict_save = st.sidebar.checkbox("Прогноз и сохранение после обучения", value=False, key="auto_predict_save_checkbox")

    if st.sidebar.button("Обучить модель"):
        df_train = st.session_state["df"]
        if df_train is None:
            st.warning("Сначала загрузите Train-данные!")
        elif tgt_col == "<нет>":
            st.error("Целевая колонка не выбрана!")
        else:
            try:
                # Удаляем старый каталог моделей
                shutil.rmtree("AutogluonModels", ignore_errors=True)

                chosen_metric_val = eval_metric if eval_metric != "auto" else None
                problem_type_val = problem_type
                fill_method_val = fill_method

                df2 = df_train.copy()
                df2 = fill_missing_values(df2, fill_method_val)

                # hyperparameters (если выбран не "* (all)")
                all_models_opt = "* (all)"
                if (len(chosen_models) == 1 and chosen_models[0] == all_models_opt) or (len(chosen_models) == 0):
                    hyperparams = None
                else:
                    no_star = [m for m in chosen_models if m != all_models_opt]
                    hyperparams = {m: {} for m in no_star}

                predictor = TabularPredictor(
                    label=tgt_col,
                    problem_type=problem_type_val if problem_type_val != "auto" else None,
                    eval_metric=chosen_metric_val,
                    path=MODEL_DIR
                ).fit(
                    train_data=df2,
                    time_limit=time_limit,
                    presets=presets,
                    hyperparameters=hyperparams
                )

                st.session_state["predictor"] = predictor
                st.success("Модель успешно обучена!")

                lb = predictor.leaderboard(df2)
                st.session_state["leaderboard"] = lb
                st.subheader("Таблица лидеров")
                st.dataframe(lb)

                fsumm = predictor.fit_summary()
                st.session_state["fit_summary"] = fsumm
                with st.expander("Резюме обучения"):
                    display_fit_summary(fsumm)

                fi = predictor.feature_importance(df2)
                st.session_state["feature_importance"] = fi
                st.subheader("Важность признаков")
                st.dataframe(fi)

                if not lb.empty:
                    best_model = lb.iloc[0]["model"]
                    best_score = lb.iloc[0]["score_val"]
                    st.info(f"Лучшая модель: {best_model}, score_val={best_score:.4f}")
                    # Проверяем ансамбль
                    ens_df = extract_ensemble_info(predictor, best_model)
                    if ens_df is not None:
                        st.session_state["ensemble_info"] = ens_df
                        st.write("### Состав ансамбля (WeightedEnsemble):")
                        st.dataframe(ens_df)

                save_model_metadata(
                    целевая_колонка=tgt_col,
                    тип_задачи=problem_type_val,
                    метрика_оценки=chosen_metric_val,
                    метод_заполнения_пропусков=fill_method_val,
                    группировочные_колонки_для_заполнения=group_cols_for_fill,
                    пресеты=presets,
                    выбранные_модели=chosen_models
                )

                # Автоматический прогноз и скачивание, если чекбокс установлен
                if auto_predict_save:
                    st.info("Делаем автопрогноз и готовим кнопку для скачивания...")
                    df_p = st.session_state.get("df_predict")
                    if df_p is None:
                        st.warning("Нет файла для прогнозирования!")
                    else:
                        try:
                            df_pred = df_p.copy()
                            df_pred = fill_missing_values(df_pred, fill_method_val)
                            preds = predictor.predict(df_pred)
                            if isinstance(preds, pd.Series):
                                preds = preds.to_frame("prediction")
                            out_df = pd.concat([df_pred.reset_index(drop=True), preds.reset_index(drop=True)], axis=1)
                            st.subheader("Автопрогноз (первые строки)")
                            st.dataframe(out_df.head())

                            # Формируем Excel в памяти
                            excel_buffer = io.BytesIO()
                            with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
                                # 1) Прогноз
                                out_df.to_excel(writer, sheet_name="РезультатыПрогноза", index=False)
                                # 2) Лидерборд
                                lb.to_excel(writer, sheet_name="ТаблицаЛидеров", index=False)
                                if not lb.empty:
                                    sheet_lb = writer.sheets["ТаблицаЛидеров"]
                                    best_idx = lb.iloc[0].name
                                    fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                                    row_excel = best_idx + 2
                                    for col_idx in range(1, lb.shape[1] + 1):
                                        cell = sheet_lb.cell(row=row_excel, column=col_idx)
                                        cell.fill = fill_green
                                # 3) Важность
                                fi.to_excel(writer, sheet_name="ВажностьПризнаков", index=True)
                                # 4) Ансамбль
                                if st.session_state["ensemble_info"] is not None:
                                    st.session_state["ensemble_info"].to_excel(writer, sheet_name="EnsembleInfo", index=False)

                            excel_buffer.seek(0)
                            st.download_button(
                                label="Скачать Excel (авто)",
                                data=excel_buffer.getvalue(),
                                file_name="results.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                            st.success("Готово! Файл Excel можно скачать.")
                        except Exception as xp:
                            st.error(f"Ошибка автопрогноза: {xp}")

            except Exception as exc:
                st.error(f"Ошибка обучения: {exc}")

    # ========== (6) Прогноз (ручной) ==========
    st.sidebar.header("6. Прогноз")
    if st.sidebar.button("Сделать прогноз"):
        predictor = st.session_state.get("predictor")
        if predictor is None:
            st.warning("Сначала обучите модель или загрузите!")
        elif tgt_col == "<нет>":
            st.error("Целевая колонка не выбрана.")
        else:
            df_p = st.session_state.get("df_predict")
            if df_p is None:
                st.error("Нет файла для прогнозирования!")
            else:
                try:
                    fill_val = st.session_state["fill_method_key"]
                    df_pred = df_p.copy()
                    df_pred = fill_missing_values(df_pred, fill_val)
                    preds = predictor.predict(df_pred)
                    if isinstance(preds, pd.Series):
                        preds = preds.to_frame("prediction")
                    out_df = pd.concat([df_pred.reset_index(drop=True), preds.reset_index(drop=True)], axis=1)
                    st.session_state["predictions"] = out_df
                    st.subheader("Предсказанные значения (первые строки)")
                    st.dataframe(out_df.head())

                    if st.session_state["ensemble_info"] is not None:
                        st.write("### Ансамбль (WeightedEnsemble):")
                        st.dataframe(st.session_state["ensemble_info"])
                except Exception as e:
                    st.error(f"Ошибка прогнозирования: {e}")

    # ========== (7) Кнопки ручного сохранения (CSV/Excel) ==========
    st.sidebar.header("7. Сохранение результатов")

    # Кнопка «Сохранить в CSV»
    if st.sidebar.button("Сохранить в CSV"):
        final_preds = st.session_state.get("predictions")
        if final_preds is None:
            st.warning("Сначала сделайте прогноз.")
        else:
            csv_data = final_preds.to_csv(index=False, encoding="utf-8-sig")
            st.download_button(
                label="Скачать CSV",
                data=csv_data,
                file_name="results.csv",
                mime="text/csv"
            )

    # Кнопка «Сохранить в Excel»
    if st.sidebar.button("Сохранить в Excel"):
        final_preds = st.session_state.get("predictions")
        lb = st.session_state.get("leaderboard")
        fi = st.session_state.get("feature_importance")
        ensemble_df = st.session_state.get("ensemble_info")

        if final_preds is None:
            st.warning("Сначала сделайте прогноз.")
        else:
            excel_buffer = io.BytesIO()
            with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
                # Прогноз
                final_preds.to_excel(writer, sheet_name="РезультатыПрогноза", index=False)

                # Лидерборд
                if lb is not None:
                    lb.to_excel(writer, sheet_name="ТаблицаЛидеров", index=False)
                    if not lb.empty:
                        sheet_lb = writer.sheets["ТаблицаЛидеров"]
                        best_idx = lb.iloc[0].name
                        fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        row_excel = best_idx + 2
                        for col_idx in range(1, lb.shape[1] + 1):
                            cell = sheet_lb.cell(row=row_excel, column=col_idx)
                            cell.fill = fill_green

                # Важность признаков
                if fi is not None:
                    fi.to_excel(writer, sheet_name="ВажностьПризнаков", index=True)

                # Ансамбль
                if ensemble_df is not None:
                    ensemble_df.to_excel(writer, sheet_name="EnsembleInfo", index=False)

            excel_buffer.seek(0)
            st.download_button(
                label="Скачать Excel",
                data=excel_buffer.getvalue(),
                file_name="results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            st.success("Файл Excel с результатами сформирован!")

    # ========== (8) Логи приложения ==========
    st.sidebar.header("8. Логи приложения")
    if st.sidebar.button("Показать логи"):
        logs_ = read_logs()
        st.subheader("Логи приложения")
        st.text(logs_)

    if st.sidebar.button("Скачать логи"):
        logs_ = read_logs()
        st.download_button(
            label="Скачать лог-файл",
            data=logs_,
            file_name="app.log",
            mime="text/plain"
        )

    # ========== (9) Скачивание модели и логов ==========
    st.sidebar.header("9. Загрузка моделей и логов")
    if st.sidebar.button("Скачать AutogluonModels + логи"):
        if not os.path.exists("AutogluonModels"):
            st.error("Папка 'AutogluonModels' не найдена (сначала обучите модель).")
        else:
            try:
                zip_buf = io.BytesIO()
                with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
                    for root, dirs, files in os.walk("AutogluonModels"):
                        for file in files:
                            fp = os.path.join(root, file)
                            arcname = os.path.relpath(fp, start="AutogluonModels")
                            zf.write(fp, arcname=arcname)

                # Добавляем лог-файл, если есть
                if os.path.exists(LOG_FILE):
                    zf.write(LOG_FILE, arcname="app.log")

                zip_buf.seek(0)
                st.download_button(
                    label="Скачать архив (модели + логи)",
                    data=zip_buf.getvalue(),
                    file_name="models_and_logs.zip",
                    mime="application/zip"
                )
            except Exception as e:
                st.error(f"Ошибка архивации: {e}")


if __name__ == "__main__":
    main()
