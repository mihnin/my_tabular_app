from fastapi import APIRouter, UploadFile, File, HTTPException, Form
from fastapi.responses import JSONResponse
import pandas as pd
import os
from fastapi import Query
from pydantic import BaseModel
import numpy as np
from concurrent.futures import ThreadPoolExecutor
import asyncio
from fastapi import Request

router = APIRouter()

@router.post("/preview-excel")
async def preview_excel(file: UploadFile = File(...)):
    try:
        filename = file.filename.lower()
        file.file.seek(0)
        if filename.endswith('.csv'):
            df = pd.read_csv(file.file, nrows=10)
            file.file.seek(0)
            # Для csv считаем строки вручную
            total_rows = sum(1 for _ in file.file) - 1  # минус строка заголовков
        else:
            df = pd.read_excel(file.file, nrows=10)
            file.file.seek(0)
            from openpyxl import load_workbook
            wb = load_workbook(file.file, read_only=True)
            ws = wb.active
            total_rows = ws.max_row - 1  # минус строка заголовков
            wb.close()
        df = df.astype(str)
        data = df.to_dict(orient="records")
        columns = list(df.columns)
        return JSONResponse({"columns": columns, "rows": data, "total_rows": total_rows})
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка обработки файла: {str(e)}")

class TimeSeriesRequest(BaseModel):
    session_id: str
    id_column: str
    ts_id: str


async def analyze_dataframe(df):
    df = df.astype(str)
    columns = list(df.columns)
    rows = df.values.tolist()
    total = len(df)
    # Пропуски по столбцам
    missing_by_column = []
    for col in columns:
        missing = int(df[col].isnull().sum() + ((df[col] == '').sum() if df[col].dtype == object else 0))
        missing_by_column.append({"column": col, "missing": missing})
    return {
        "columns": columns,
        "rows": rows[:10],
        "total": total,
        "missing_by_column": missing_by_column
    }


@router.post("/analyze-tabular")
async def analyze_tabular(
    train_file: UploadFile = File(None),
    predict_file: UploadFile = File(None),
    session_id: str = Form(None)
):
    """
    Анализирует train и predict файлы (или parquet по session_id), возвращает оба результата.
    Если session_id передан — ищет train.parquet и prediction.parquet в папке сессии.
    """
    loop = asyncio.get_event_loop()
    try:
        def load_df(file, parquet_name):
            if session_id:
                session_dir = os.path.join("training_sessions", session_id)
                parquet_path = os.path.join(session_dir, parquet_name)
                if not os.path.exists(parquet_path):
                    return None
                return pd.read_parquet(parquet_path)
            elif file and file.filename:
                filename = file.filename.lower()
                if filename.endswith('.csv'):
                    return pd.read_csv(file.file)
                else:
                    return pd.read_excel(file.file)
            else:
                return None

        with ThreadPoolExecutor() as pool:
            train_df = await loop.run_in_executor(pool, lambda: load_df(train_file, "train.parquet"))
            train_result = await analyze_dataframe(train_df) if train_df is not None else None
            predict_df = await loop.run_in_executor(pool, lambda: load_df(predict_file, "prediction.parquet"))
            predict_result = await analyze_dataframe(predict_df) if predict_df is not None else None
        return JSONResponse({
            "train": train_result,
            "predict": predict_result
        })
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка анализа данных: {str(e)}")
